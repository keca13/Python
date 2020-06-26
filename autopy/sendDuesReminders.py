#! python3
# sendDuesReminders.py - Wysyłanie wiadomości e-mail na podstawie
# zapisanych w arkuszu kalkulacyjnym informacji o płatnościach.

import openpyxl, smtplib, sys

# Otworzenie arkusza kalkulacyjnego i pobranie informacji o ostatnich płatnościach.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
# Sprawdzenie stanu płatności poszczególnych członków klubu.
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'zapłacono':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Zalogowanie do konta poczty elektronicznej.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('mój_adres_email@gmail.com', sys.argv[1])

# Wysłanie wiadomości e-mail z przypomnieniem o składce.
for name, email in unpaidMembers.items():
    body = 'Subject: Zaległa składka za %s.\nWitaj, %s!\nPrawdopodobnie masz niezapłaconą składkę za miesiąc %s. Proszę o jak najszybsze uregulowanie należności. Dziękuję!' % (latestMonth, name, latestMonth)
    print('Wysyłanie wiadomości e-mail na adres %s...' % email)
    sendmailStatus = smtpObj.sendmail('mój_adres_email@gmail.com', email, body)

    if sendmailStatus != {}:
        print('Wystąpił problem podczas wysyłania wiadomości e-mail na adres %s: %s' % (email, sendmailStatus))
smtpObj.quit()
