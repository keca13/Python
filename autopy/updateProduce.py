#! python3
# updateProduce.py - Korekta ceny produktu w arkuszu kalkulacyjnym zawierającym dane sprzedaży.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# Poszczególne produkty i ich uaktualnione ceny.
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Iteracja przez wiersze i uaktualnienie cen.
for rowNum in range(2, sheet.get_highest_row()): # Pominięcie pierwszego wiersza.
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
